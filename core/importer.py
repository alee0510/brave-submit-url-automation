import csv
import os
import re
from urllib.parse import urlparse

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

URL_PATTERN = re.compile(r"^https?://", re.IGNORECASE)

def _is_valid_url(value):
    if not value:
        return False
    value = value.strip()
    if not URL_PATTERN.match(value):
        return False
    parsed = urlparse(value)
    return bool(parsed.netloc)

def _read_csv(path):
    with open(path, "r", newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))

    # Skip a header row if the first cell isn't itself a URL
    if rows and not _is_valid_url(rows[0][0] if rows[0] else ""):
        rows = rows[1:]

    return [row[0].strip() for row in rows if row and row[0].strip()]

def _read_xlsx(path):
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to import .xlsx files (pip install openpyxl)")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    values = [str(row[0]).strip() for row in ws.iter_rows(values_only=True) if row and row[0]]

    if values and not _is_valid_url(values[0]):
        values = values[1:]

    return values

def import_urls(queue_manager, file_path):
    """
    Reads URLs from a .csv or .xlsx file (first column) and adds valid,
    de-duplicated ones to the queue. Returns (added, invalid, duplicate) counts.
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        candidates = _read_csv(file_path)
    elif ext == ".xlsx":
        candidates = _read_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file type: {ext} (use .csv or .xlsx)")

    added, invalid, duplicate = 0, 0, 0

    for raw in candidates:
        url = raw.strip()
        if not _is_valid_url(url):
            invalid += 1
            continue
        if queue_manager.add_url(url):
            added += 1
        else:
            duplicate += 1

    if added:
        queue_manager.save()

    return added, invalid, duplicate